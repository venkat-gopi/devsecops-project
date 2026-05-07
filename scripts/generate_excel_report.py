#!/usr/bin/env python3

import os
import json
import argparse
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLORS = {
    "CRITICAL": {"fill": "FFDC2626", "font": "FFFFFFFF"},
    "HIGH": {"fill": "FFEA580C", "font": "FFFFFFFF"},
    "MEDIUM": {"fill": "FFFEF08A", "font": "FF78350F"},
    "LOW": {"fill": "FFD1FAE5", "font": "FF14532D"},
    "INFO": {"fill": "FFE0F2FE", "font": "FF0C4A6E"},
    "PASS": {"fill": "FFD1FAE5", "font": "FF14532D"},
    "BLOCKED": {"fill": "FFDC2626", "font": "FFFFFFFF"},
    "HEADER": {"fill": "FF111827", "font": "FFFFFFFF"},
    "SUBHEADER": {"fill": "FF374151", "font": "FFFFFFFF"},
    "ALT_ROW": {"fill": "FFF9FAFB", "font": "FF111827"},
    "WHITE": {"fill": "FFFFFFFF", "font": "FF111827"},
}


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def make_font(hex_color, bold=False, size=11):
    return Font(color=hex_color, bold=bold, size=size, name="Calibri")


def make_border():
    side = Side(style="thin", color="FFD1D5DB")
    return Border(left=side, right=side, top=side, bottom=side)


def style_cell(cell, style_name="WHITE", bold=False, center=False):
    style = COLORS.get(str(style_name).upper(), COLORS["WHITE"])
    cell.fill = make_fill(style["fill"])
    cell.font = make_font(style["font"], bold=bold)
    cell.border = make_border()
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )


def write_header_row(ws, headers, row=1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        style_cell(cell, "HEADER", bold=True, center=True)
    ws.row_dimensions[row].height = 24


def auto_column_width(ws, min_width=14, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, min_width), max_width)


def load_json(path):
    try:
        if not os.path.exists(path):
            return None
        with open(path, "r", encoding="utf-8") as f:
            content = f.read().strip()
            if not content:
                return None
            return json.loads(content)
    except Exception:
        return None


def short(value, limit=160):
    if value is None:
        return "N/A"
    value = str(value).replace("\n", " ").strip()
    return value[:limit]


def add_finding(findings, layer, tool, severity, category, file_path, package, issue, fix, evidence):
    findings.append({
        "Layer": layer,
        "Tool": tool,
        "Severity": str(severity or "INFO").upper(),
        "Category": category or "N/A",
        "File/Resource": file_path or "N/A",
        "Package": package or "N/A",
        "Issue": issue or "N/A",
        "Fix/Recommendation": fix or "N/A",
        "Evidence": evidence or "N/A",
    })


def collect_gitleaks_findings(reports_dir):
    findings = []
    data = load_json(os.path.join(reports_dir, "gitleaks-report.json"))

    if isinstance(data, list):
        for leak in data:
            if not isinstance(leak, dict):
                continue

            add_finding(
                findings,
                "Layer 1 - Secret Scanning",
                "GitLeaks",
                "CRITICAL",
                "Hardcoded Secret",
                leak.get("File"),
                leak.get("RuleID"),
                f"Secret detected in {leak.get('File', 'unknown file')}",
                "Remove the secret and rotate the exposed credential immediately.",
                f"Line: {leak.get('StartLine', 'N/A')} | Rule: {leak.get('RuleID', 'N/A')}",
            )

    return findings


def collect_dependency_findings(reports_dir):
    findings = []

    npm = load_json(os.path.join(reports_dir, "npm-audit.json"))
    if isinstance(npm, dict):
        vulnerabilities = npm.get("vulnerabilities", {})
        if isinstance(vulnerabilities, dict):
            for pkg, info in vulnerabilities.items():
                if not isinstance(info, dict):
                    continue

                fix_available = info.get("fixAvailable", "N/A")
                if isinstance(fix_available, dict):
                    fix = f"Upgrade to {fix_available.get('version', 'recommended fixed version')}"
                elif fix_available is True:
                    fix = "Run npm audit fix"
                else:
                    fix = "No automatic fix available"

                add_finding(
                    findings,
                    "Layer 3 - Dependency Scanning",
                    "npm audit",
                    info.get("severity", "INFO"),
                    "Vulnerable Dependency",
                    "package.json/package-lock.json",
                    pkg,
                    info.get("title") or f"Vulnerability found in {pkg}",
                    fix,
                    f"Range: {info.get('range', 'N/A')}",
                )

    snyk = load_json(os.path.join(reports_dir, "snyk-report.json"))
    if isinstance(snyk, dict):
        for vuln in snyk.get("vulnerabilities", []):
            if not isinstance(vuln, dict):
                continue

            cves = "N/A"
            identifiers = vuln.get("identifiers", {})
            if isinstance(identifiers, dict):
                cve_list = identifiers.get("CVE", [])
                if isinstance(cve_list, list) and cve_list:
                    cves = ", ".join(cve_list)

            fixed_in = vuln.get("fixedIn", [])
            if isinstance(fixed_in, list) and fixed_in:
                fix = "Upgrade to " + ", ".join(fixed_in)
            else:
                fix = "No fixed version available"

            add_finding(
                findings,
                "Layer 3B - Snyk Dependency Scan",
                "Snyk",
                vuln.get("severity", "INFO"),
                "Open Source Vulnerability",
                "dependency manifest",
                vuln.get("packageName"),
                vuln.get("title"),
                fix,
                f"CVE: {cves} | Version: {vuln.get('version', 'N/A')}",
            )

    safety = load_json(os.path.join(reports_dir, "safety-report.json"))
    if isinstance(safety, list):
        for vuln in safety:
            if not isinstance(vuln, dict):
                continue

            add_finding(
                findings,
                "Layer 3 - Dependency Scanning",
                "Safety",
                vuln.get("severity", "HIGH"),
                "Python Dependency Vulnerability",
                "requirements.txt",
                vuln.get("package_name") or vuln.get("package"),
                vuln.get("advisory") or vuln.get("vulnerability_id"),
                "Upgrade vulnerable Python package.",
                f"Installed version: {vuln.get('analyzed_version', 'N/A')}",
            )

    return findings


def collect_trivy_findings(reports_dir):
    findings = []
    trivy = load_json(os.path.join(reports_dir, "trivy-report.json"))

    if isinstance(trivy, dict):
        for result in trivy.get("Results", []):
            if not isinstance(result, dict):
                continue

            target = result.get("Target", "container image")

            for vuln in result.get("Vulnerabilities") or []:
                if not isinstance(vuln, dict):
                    continue

                add_finding(
                    findings,
                    "Layer 4 - Container Image Scanning",
                    "Trivy",
                    vuln.get("Severity", "INFO"),
                    "Container CVE",
                    target,
                    vuln.get("PkgName"),
                    vuln.get("Title") or vuln.get("VulnerabilityID"),
                    f"Fixed version: {vuln.get('FixedVersion', 'No fix available')}",
                    f"CVE: {vuln.get('VulnerabilityID', 'N/A')} | Installed: {vuln.get('InstalledVersion', 'N/A')}",
                )

    return findings


def collect_checkov_findings(reports_dir):
    findings = []
    checkov = load_json(os.path.join(reports_dir, "checkov-report.json"))

    def parse_block(block):
        if not isinstance(block, dict):
            return

        check_type = block.get("check_type") or block.get("framework") or "IaC"
        results = block.get("results", {})
        if not isinstance(results, dict):
            return

        failed_checks = results.get("failed_checks", [])
        if not isinstance(failed_checks, list):
            return

        for check in failed_checks:
            if not isinstance(check, dict):
                continue

            file_path = check.get("file_path") or check.get("file_abs_path") or "N/A"
            resource = check.get("resource") or "N/A"
            check_name = check.get("check_name") or "IaC misconfiguration detected"
            guideline = check.get("guideline") or "Review and fix the misconfiguration as per cloud security best practices."

            add_finding(
                findings,
                "Layer 5 - Infrastructure as Code Scanning",
                "Checkov",
                "HIGH",
                f"IaC Misconfiguration - {check_type}",
                file_path,
                resource,
                check_name,
                short(guideline, 220),
                f"Check ID: {check.get('check_id', 'N/A')} | Resource: {resource}",
            )

    if isinstance(checkov, list):
        for block in checkov:
            parse_block(block)
    elif isinstance(checkov, dict):
        parse_block(checkov)

    return findings


def collect_opa_findings(reports_dir):
    findings = []
    opa = load_json(os.path.join(reports_dir, "opa-result.json"))

    if isinstance(opa, dict):
        result = opa.get("result", [])
        if isinstance(result, list):
            for result_item in result:
                expressions = result_item.get("expressions", []) if isinstance(result_item, dict) else []
                for expr in expressions:
                    value = expr.get("value") if isinstance(expr, dict) else None

                    if isinstance(value, list):
                        for msg in value:
                            add_finding(
                                findings,
                                "Layer 6 - OPA Policy Gate",
                                "OPA",
                                "BLOCKED",
                                "Policy Violation",
                                "policy/deployment_policy.rego",
                                "Deployment Gate",
                                str(msg),
                                "Manual approval is required before deployment.",
                                "OPA deny rule triggered",
                            )

                    elif isinstance(value, dict):
                        for _, msg in value.items():
                            add_finding(
                                findings,
                                "Layer 6 - OPA Policy Gate",
                                "OPA",
                                "BLOCKED",
                                "Policy Violation",
                                "policy/deployment_policy.rego",
                                "Deployment Gate",
                                str(msg),
                                "Manual approval is required before deployment.",
                                "OPA deny rule triggered",
                            )

    return findings


def collect_all_findings(reports_dir):
    findings = []
    findings.extend(collect_gitleaks_findings(reports_dir))
    findings.extend(collect_dependency_findings(reports_dir))
    findings.extend(collect_trivy_findings(reports_dir))
    findings.extend(collect_checkov_findings(reports_dir))
    findings.extend(collect_opa_findings(reports_dir))
    return findings


def calculate_counts(findings):
    counts = {
        "critical": 0,
        "high": 0,
        "medium": 0,
        "low": 0,
        "info": 0,
        "blocked": 0,
        "total": 0,
    }

    for finding in findings:
        severity = finding.get("Severity", "INFO").upper()

        if severity == "CRITICAL":
            counts["critical"] += 1
            counts["total"] += 1
        elif severity == "HIGH":
            counts["high"] += 1
            counts["total"] += 1
        elif severity == "MEDIUM":
            counts["medium"] += 1
            counts["total"] += 1
        elif severity == "LOW":
            counts["low"] += 1
            counts["total"] += 1
        elif severity == "BLOCKED":
            counts["blocked"] += 1
        else:
            counts["info"] += 1

    return counts


def build_summary_sheet(ws, app_name, run_id, findings):
    counts = calculate_counts(findings)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22

    ws.merge_cells("A1:B1")
    ws["A1"].value = f"Security Scan Report — {app_name}"
    ws["A1"].font = Font(bold=True, size=15, color="FF111827", name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    ws["A2"].value = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(size=10, color="FF6B7280", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center")

    write_header_row(ws, ["Category", "Count"], row=4)

    summary_rows = [
        ("Critical Vulnerabilities", counts["critical"], "CRITICAL"),
        ("High Vulnerabilities", counts["high"], "HIGH"),
        ("Medium Vulnerabilities", counts["medium"], "MEDIUM"),
        ("Low Vulnerabilities", counts["low"], "LOW"),
        ("OPA Policy Deny Messages", counts["blocked"], "BLOCKED" if counts["blocked"] else "PASS"),
        ("Total Security Findings", counts["total"], "SUBHEADER"),
    ]

    for row_num, (label, count, style_name) in enumerate(summary_rows, start=5):
        ws.cell(row=row_num, column=1, value=label)
        ws.cell(row=row_num, column=2, value=count)
        style_cell(ws.cell(row=row_num, column=1), style_name)
        style_cell(ws.cell(row=row_num, column=2), style_name, center=True)

    info_start = 13
    ws.cell(row=info_start, column=1, value="Pipeline Run ID")
    ws.cell(row=info_start, column=2, value=run_id)
    ws.cell(row=info_start + 1, column=1, value="Report Format")
    ws.cell(row=info_start + 1, column=2, value="Excel (.xlsx) - Detailed DevSecOps Report")

    for r in range(info_start, info_start + 2):
        style_cell(ws.cell(row=r, column=1), "ALT_ROW", bold=True)
        style_cell(ws.cell(row=r, column=2), "WHITE")

    ws.cell(row=17, column=1, value="Customer Note")
    ws.cell(
        row=17,
        column=2,
        value="Open the 'All Findings' sheet to view the actual vulnerabilities, affected files/packages, and recommendations.",
    )
    style_cell(ws.cell(row=17, column=1), "INFO", bold=True)
    style_cell(ws.cell(row=17, column=2), "INFO")

    auto_column_width(ws)


def write_findings_sheet(ws, findings, title_if_empty):
    headers = [
        "Layer",
        "Tool",
        "Severity",
        "Category",
        "File/Resource",
        "Package",
        "Issue",
        "Fix/Recommendation",
        "Evidence",
    ]

    write_header_row(ws, headers)

    if not findings:
        ws.append([title_if_empty, "", "", "", "", "", "", "", ""])
        style_cell(ws.cell(row=2, column=1), "PASS")
        auto_column_width(ws)
        return

    for idx, finding in enumerate(findings, start=2):
        severity = finding.get("Severity", "INFO").upper()

        values = [finding.get(h, "N/A") for h in headers]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)

            if col == 3:
                style_cell(cell, severity, bold=True, center=True)
            else:
                style_cell(cell, "ALT_ROW" if idx % 2 == 0 else "WHITE")

        ws.row_dimensions[idx].height = 36

    auto_column_width(ws)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reports-dir", required=True)
    parser.add_argument("--app-name", required=True)
    parser.add_argument("--run-id", default="local")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    os.makedirs(os.path.dirname(args.output), exist_ok=True)

    all_findings = collect_all_findings(args.reports_dir)

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    build_summary_sheet(ws_summary, args.app_name, args.run_id, all_findings)

    ws_all = wb.create_sheet("All Findings")
    write_findings_sheet(ws_all, all_findings, "No security findings detected")

    ws_gitleaks = wb.create_sheet("GitLeaks Secrets")
    write_findings_sheet(
        ws_gitleaks,
        collect_gitleaks_findings(args.reports_dir),
        "No secrets detected by GitLeaks",
    )

    ws_deps = wb.create_sheet("Dependencies")
    write_findings_sheet(
        ws_deps,
        collect_dependency_findings(args.reports_dir),
        "No dependency vulnerabilities detected",
    )

    ws_trivy = wb.create_sheet("Container Trivy")
    write_findings_sheet(
        ws_trivy,
        collect_trivy_findings(args.reports_dir),
        "No container vulnerabilities detected by Trivy",
    )

    ws_checkov = wb.create_sheet("Checkov IaC")
    write_findings_sheet(
        ws_checkov,
        collect_checkov_findings(args.reports_dir),
        "No IaC misconfigurations detected by Checkov",
    )

    ws_opa = wb.create_sheet("OPA Policy Gate")
    write_findings_sheet(
        ws_opa,
        collect_opa_findings(args.reports_dir),
        "OPA policy passed. No deny messages found.",
    )

    wb.save(args.output)

    counts = calculate_counts(all_findings)

    print(f"Excel report saved: {args.output}")
    print(f"Critical findings: {counts['critical']}")
    print(f"High findings: {counts['high']}")
    print(f"Medium findings: {counts['medium']}")
    print(f"Low findings: {counts['low']}")
    print(f"OPA deny messages: {counts['blocked']}")
    print(f"Total security findings: {counts['total']}")


if __name__ == "__main__":
    main()
