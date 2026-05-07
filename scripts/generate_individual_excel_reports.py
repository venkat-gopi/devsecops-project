#!/usr/bin/env python3

import os
import json
import argparse
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLORS = {
    "CRITICAL": {"fill": "FFDC2626", "font": "FFFFFFFF"},
    "HIGH": {"fill": "FFEA580C", "font": "FFFFFFFF"},
    "MEDIUM": {"fill": "FFFEF08A", "font": "FF78350F"},
    "LOW": {"fill": "FFD1FAE5", "font": "FF14532D"},
    "INFO": {"fill": "FFE0F2FE", "font": "FF0C4A6E"},
    "PASS": {"fill": "FFD1FAE5", "font": "FF14532D"},
    "BLOCKED": {"fill": "FFDC2626", "font": "FFFFFFFF"},
    "HEADER": {"fill": "FF111827", "font": "FFFFFFFF"},
    "ALT": {"fill": "FFF9FAFB", "font": "FF111827"},
    "WHITE": {"fill": "FFFFFFFF", "font": "FF111827"},
}


def load_json(path):
    try:
        if not os.path.exists(path):
            return None

        with open(path, "r", encoding="utf-8") as f:
            content = f.read().strip()

        if not content:
            return None

        return json.loads(content)

    except Exception:
        return None


def style_cell(cell, style="WHITE", bold=False, center=False):
    config = COLORS.get(str(style).upper(), COLORS["WHITE"])

    cell.fill = PatternFill("solid", fgColor=config["fill"])
    cell.font = Font(color=config["font"], bold=bold, name="Calibri", size=11)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )

    side = Side(style="thin", color="FFD1D5DB")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 70)


def write_title(ws, title, app_name):
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=15, color="FF111827", name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Application: {app_name} | "
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    )
    ws["A2"].font = Font(size=10, color="FF6B7280", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center")


def write_table(ws, headers, rows, empty_message):
    start_row = 4

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        style_cell(cell, "HEADER", bold=True, center=True)

    if not rows:
        ws.cell(row=start_row + 1, column=1, value=empty_message)
        style_cell(ws.cell(row=start_row + 1, column=1), "PASS", bold=True)
        auto_width(ws)
        return

    for r_idx, row in enumerate(rows, start_row + 1):
        severity = str(row.get("Severity", "INFO")).upper()

        for c_idx, header in enumerate(headers, 1):
            value = row.get(header, "N/A")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if header == "Severity":
                style_cell(cell, severity, bold=True, center=True)
            else:
                style_cell(cell, "ALT" if r_idx % 2 == 0 else "WHITE")

        ws.row_dimensions[r_idx].height = 35

    auto_width(ws)


def save_workbook(output_path, sheet_title, app_name, headers, rows, empty_message):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    write_title(ws, f"{sheet_title} Report", app_name)
    write_table(ws, headers, rows, empty_message)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    print(f"Generated Excel report: {output_path}")


# ---------------- GITLEAKS ----------------

def generate_gitleaks(reports_dir, output_dir, app_name):
    data = load_json(os.path.join(reports_dir, "gitleaks-report.json"))
    rows = []

    if isinstance(data, list):
        for leak in data:
            if not isinstance(leak, dict):
                continue

            rows.append({
                "Tool": "GitLeaks",
                "Severity": "CRITICAL",
                "Rule": leak.get("RuleID", "N/A"),
                "File": leak.get("File", "N/A"),
                "Line": leak.get("StartLine", "N/A"),
                "Issue": "Hardcoded secret detected",
                "Recommendation": "Remove secret and rotate the exposed credential immediately",
                "Evidence": str(leak.get("Match", "N/A"))[:180],
            })

    headers = [
        "Tool",
        "Severity",
        "Rule",
        "File",
        "Line",
        "Issue",
        "Recommendation",
        "Evidence",
    ]

    save_workbook(
        os.path.join(output_dir, "gitleaks-report.xlsx"),
        "GitLeaks Secrets",
        app_name,
        headers,
        rows,
        "No secrets detected by GitLeaks",
    )


# ---------------- SAST / SONARCLOUD ----------------

def generate_sast(reports_dir, output_dir, app_name):
    data = load_json(os.path.join(reports_dir, "sonar-report.json"))
    rows = []

    if isinstance(data, dict):
        status = data.get("status", "completed")
        message = data.get(
            "message",
            "SonarCloud scan completed. Review SonarCloud dashboard for detailed SAST findings.",
        )

        rows.append({
            "Tool": "SonarCloud",
            "Severity": "INFO" if status != "failed" else "HIGH",
            "Category": "SAST",
            "Status": status,
            "Issue": message,
            "File": "app/",
            "Recommendation": "Review SonarCloud dashboard for bugs, vulnerabilities, and code smells.",
            "Evidence": json.dumps(data)[:250],
        })

    else:
        rows.append({
            "Tool": "SonarCloud",
            "Severity": "INFO",
            "Category": "SAST",
            "Status": "completed",
            "Issue": "SonarCloud scan executed. Check SonarCloud dashboard for detailed issues.",
            "File": "app/",
            "Recommendation": "Review SonarCloud dashboard.",
            "Evidence": "N/A",
        })

    headers = [
        "Tool",
        "Severity",
        "Category",
        "Status",
        "Issue",
        "File",
        "Recommendation",
        "Evidence",
    ]

    save_workbook(
        os.path.join(output_dir, "sast-sonar-report.xlsx"),
        "SAST SonarCloud",
        app_name,
        headers,
        rows,
        "No SAST issues exported",
    )


# ---------------- DEPENDENCY: NPM + SAFETY ----------------

def generate_dependencies(reports_dir, output_dir, app_name):
    rows = []

    npm = load_json(os.path.join(reports_dir, "npm-audit.json"))

    if isinstance(npm, dict):
        vulns = npm.get("vulnerabilities", {})

        if isinstance(vulns, dict):
            for package, info in vulns.items():
                if not isinstance(info, dict):
                    continue

                fix_available = info.get("fixAvailable", "N/A")

                if isinstance(fix_available, dict):
                    fix = f"Upgrade to {fix_available.get('version', 'recommended fixed version')}"
                elif fix_available is True:
                    fix = "Run npm audit fix"
                else:
                    fix = "No automatic fix available"

                rows.append({
                    "Tool": "npm audit",
                    "Severity": str(info.get("severity", "INFO")).upper(),
                    "Package": package,
                    "Version/Range": info.get("range", "N/A"),
                    "Issue": info.get("title", f"Vulnerability in {package}"),
                    "Fix": fix,
                    "File": "package.json/package-lock.json",
                    "Evidence": str(info)[:250],
                })

    safety = load_json(os.path.join(reports_dir, "safety-report.json"))

    if isinstance(safety, list):
        for vuln in safety:
            if not isinstance(vuln, dict):
                continue

            rows.append({
                "Tool": "Safety",
                "Severity": str(vuln.get("severity", "HIGH")).upper(),
                "Package": vuln.get("package_name") or vuln.get("package", "N/A"),
                "Version/Range": vuln.get("analyzed_version", "N/A"),
                "Issue": vuln.get("advisory") or vuln.get("vulnerability_id", "Python dependency vulnerability"),
                "Fix": "Upgrade vulnerable Python dependency",
                "File": "requirements.txt",
                "Evidence": str(vuln)[:250],
            })

    headers = [
        "Tool",
        "Severity",
        "Package",
        "Version/Range",
        "Issue",
        "Fix",
        "File",
        "Evidence",
    ]

    save_workbook(
        os.path.join(output_dir, "dependency-report.xlsx"),
        "Dependency Scan",
        app_name,
        headers,
        rows,
        "No dependency vulnerabilities detected",
    )


# ---------------- SNYK ----------------

def generate_snyk(reports_dir, output_dir, app_name):
    data = load_json(os.path.join(reports_dir, "snyk-report.json"))
    rows = []

    if isinstance(data, dict):
        vulns = data.get("vulnerabilities", [])

        if isinstance(vulns, list):
            for vuln in vulns:
                if not isinstance(vuln, dict):
                    continue

                identifiers = vuln.get("identifiers", {})
                cves = "N/A"

                if isinstance(identifiers, dict):
                    cve_list = identifiers.get("CVE", [])
                    if isinstance(cve_list, list) and cve_list:
                        cves = ", ".join(cve_list)

                fixed_in = vuln.get("fixedIn", [])
                fix = ", ".join(fixed_in) if isinstance(fixed_in, list) and fixed_in else "No fix available"

                rows.append({
                    "Tool": "Snyk",
                    "Severity": str(vuln.get("severity", "INFO")).upper(),
                    "Package": vuln.get("packageName", "N/A"),
                    "Version": vuln.get("version", "N/A"),
                    "Issue": vuln.get("title", "N/A"),
                    "CVE": cves,
                    "Fix": fix,
                    "Evidence": str(vuln)[:250],
                })

    headers = [
        "Tool",
        "Severity",
        "Package",
        "Version",
        "Issue",
        "CVE",
        "Fix",
        "Evidence",
    ]

    save_workbook(
        os.path.join(output_dir, "snyk-report.xlsx"),
        "Snyk Scan",
        app_name,
        headers,
        rows,
        "No Snyk vulnerabilities detected or Snyk scan skipped",
    )


# ---------------- TRIVY ----------------

def generate_trivy(reports_dir, output_dir, app_name):
    data = load_json(os.path.join(reports_dir, "trivy-report.json"))
    rows = []

    if isinstance(data, dict):
        for result in data.get("Results", []):
            if not isinstance(result, dict):
                continue

            target = result.get("Target", "container image")
            vulns = result.get("Vulnerabilities") or []

            for vuln in vulns:
                if not isinstance(vuln, dict):
                    continue

                rows.append({
                    "Tool": "Trivy",
                    "Severity": str(vuln.get("Severity", "INFO")).upper(),
                    "CVE": vuln.get("VulnerabilityID", "N/A"),
                    "Package": vuln.get("PkgName", "N/A"),
                    "Installed Version": vuln.get("InstalledVersion", "N/A"),
                    "Fixed Version": vuln.get("FixedVersion", "No fix available"),
                    "Target": target,
                    "Issue": vuln.get("Title", "N/A"),
                })

    headers = [
        "Tool",
        "Severity",
        "CVE",
        "Package",
        "Installed Version",
        "Fixed Version",
        "Target",
        "Issue",
    ]

    save_workbook(
        os.path.join(output_dir, "trivy-container-report.xlsx"),
        "Trivy Container",
        app_name,
        headers,
        rows,
        "No container vulnerabilities detected by Trivy",
    )


# ---------------- CHECKOV ----------------

def generate_checkov(reports_dir, output_dir, app_name):
    data = load_json(os.path.join(reports_dir, "checkov-report.json"))
    rows = []

    def parse_block(block):
        if not isinstance(block, dict):
            return

        check_type = block.get("check_type") or block.get("framework") or "IaC"
        results = block.get("results", {})

        if not isinstance(results, dict):
            return

        failed_checks = results.get("failed_checks", [])

        if not isinstance(failed_checks, list):
            return

        for check in failed_checks:
            if not isinstance(check, dict):
                continue

            rows.append({
                "Tool": "Checkov",
                "Severity": "HIGH",
                "Framework": check_type,
                "Check ID": check.get("check_id", "N/A"),
                "File": check.get("file_path") or check.get("file_abs_path", "N/A"),
                "Resource": check.get("resource", "N/A"),
                "Issue": check.get("check_name", "IaC misconfiguration detected"),
                "Guideline": str(check.get("guideline") or "Review security best practices")[:250],
            })

    if isinstance(data, list):
        for block in data:
            parse_block(block)

    elif isinstance(data, dict):
        parse_block(data)

    headers = [
        "Tool",
        "Severity",
        "Framework",
        "Check ID",
        "File",
        "Resource",
        "Issue",
        "Guideline",
    ]

    save_workbook(
        os.path.join(output_dir, "checkov-iac-report.xlsx"),
        "Checkov IaC",
        app_name,
        headers,
        rows,
        "No IaC misconfigurations detected by Checkov",
    )


# ---------------- OPA ----------------

def generate_opa(reports_dir, output_dir, app_name):
    data = load_json(os.path.join(reports_dir, "opa-result.json"))
    rows = []

    if isinstance(data, dict):
        result = data.get("result", [])

        if isinstance(result, list):
            for item in result:
                if not isinstance(item, dict):
                    continue

                expressions = item.get("expressions", [])

                for expr in expressions:
                    if not isinstance(expr, dict):
                        continue

                    value = expr.get("value")

                    if isinstance(value, list):
                        for msg in value:
                            rows.append({
                                "Tool": "OPA",
                                "Severity": "BLOCKED",
                                "Policy": "deployment_policy.rego",
                                "Decision": "DENY",
                                "Message": str(msg),
                                "Recommendation": "Manual approval or vulnerability remediation required",
                                "Input": "scan-summary.json",
                                "Evidence": "OPA deny rule triggered",
                            })

                    elif isinstance(value, dict):
                        for _, msg in value.items():
                            rows.append({
                                "Tool": "OPA",
                                "Severity": "BLOCKED",
                                "Policy": "deployment_policy.rego",
                                "Decision": "DENY",
                                "Message": str(msg),
                                "Recommendation": "Manual approval or vulnerability remediation required",
                                "Input": "scan-summary.json",
                                "Evidence": "OPA deny rule triggered",
                            })

    headers = [
        "Tool",
        "Severity",
        "Policy",
        "Decision",
        "Message",
        "Recommendation",
        "Input",
        "Evidence",
    ]

    save_workbook(
        os.path.join(output_dir, "opa-policy-report.xlsx"),
        "OPA Policy Gate",
        app_name,
        headers,
        rows,
        "OPA policy passed. No deny messages found.",
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reports-dir", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--app-name", required=True)

    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    generate_gitleaks(args.reports_dir, args.output_dir, args.app_name)
    generate_sast(args.reports_dir, args.output_dir, args.app_name)
    generate_dependencies(args.reports_dir, args.output_dir, args.app_name)
    generate_snyk(args.reports_dir, args.output_dir, args.app_name)
    generate_trivy(args.reports_dir, args.output_dir, args.app_name)
    generate_checkov(args.reports_dir, args.output_dir, args.app_name)
    generate_opa(args.reports_dir, args.output_dir, args.app_name)

    print("All individual Excel reports generated successfully.")


if __name__ == "__main__":
    main()
